from logging import info
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from pathlib import Path

app = FastAPI(title="Cost ManagementAPI")
ARQUIVO = Path("storage.xlsx")

# --------- MODEL ----------
class Gasto(BaseModel):
    data: str
    classificacao: str
    gasto: float

# --------- HELPERS ----------
def salvar_em_planilha(gasto: Gasto):
    if not ARQUIVO.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"
        ws.append(["Data", "Classificacao", "Gasto"])
        wb.save(ARQUIVO)
    wb = load_workbook(ARQUIVO)
    ws = wb.active

    ws.append([gasto.data, gasto.classificacao, gasto.gasto])
    wb.save(ARQUIVO)

# --------- ENDPOINTS ----------
@app.post("/webhook")
def add_cost(info: Gasto):
    try:
        salvar_em_planilha(info)
        return {"message": "Cost item added successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))